import io
from datetime import date, datetime

import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from src.service_layer.unit_of_work import AbstractUnitOfWork

_SHEET_NAME = "Relatório de Ponto"
_COLUMN_WIDTHS = [12, 12, 12, 12, 12, 12, 15, 30]


def generate_excel_report(uow: AbstractUnitOfWork, user_id: int, start_date: date = None, end_date: date = None) -> io.BytesIO:
        with uow:
                user = _require_user(uow, user_id)
                rows = _report_rows(user, start_date, end_date)
                period = _period_label(start_date, end_date)
                return _build_workbook(user, rows, period)


def _require_user(uow, user_id):
        user = uow.users.get_user_by_id(user_id)
        if not user:
                raise ValueError("Usuário não encontrado.")
        return user


def _report_rows(user, start_date, end_date):
        return [_row_data(entry) for entry in _filter_entries(user.time_entries, start_date, end_date)]


def _filter_entries(entries, start_date, end_date):
        if start_date:
                entries = [entry for entry in entries if entry.entry_date >= start_date]
        if end_date:
                entries = [entry for entry in entries if entry.entry_date <= end_date]
        return sorted(entries, key=lambda entry: entry.entry_date)


def _row_data(entry):
        return {
                "Data": entry.entry_date.strftime("%d/%m/%Y"),
                "Chegada": _fmt_time(entry.arrival),
                "Saída Almoço": _fmt_time(entry.lunch_start),
                "Volta Almoço": _fmt_time(entry.lunch_end),
                "Fim Jornada": _fmt_time(entry.departure),
                "Horas Trab.": f"{entry.worked_minutes // 60:02d}:{entry.worked_minutes % 60:02d}",
                "Status": entry.status.value,
                "Observações": entry.notes or "",
        }


def _fmt_time(value):
        return value.strftime("%H:%M:%S") if value else "-"


def _period_label(start_date, end_date):
        if start_date and end_date:
                return f"{start_date.strftime('%d/%m/%Y')} até {end_date.strftime('%d/%m/%Y')}"
        if start_date:
                return f"A partir de {start_date.strftime('%d/%m/%Y')}"
        if end_date:
                return f"Até {end_date.strftime('%d/%m/%Y')}"
        return "Todo o período"


def _build_workbook(user, rows, period):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
                pd.DataFrame(rows).to_excel(writer, index=False, sheet_name=_SHEET_NAME, startrow=7)
                sheet = writer.book[_SHEET_NAME]
                styles = _report_styles()
                _write_header(sheet, period, user, styles)
                _style_table(sheet, len(rows), styles)
                _write_signatures(sheet, len(rows))
        output.seek(0)
        return output


def _report_styles():
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        return {
                "title": Font(name="Arial", size=14, bold=True),
                "header": Font(name="Arial", size=10, bold=True),
                "label": Font(name="Arial", size=9, bold=True),
                "value": Font(name="Arial", size=9),
                "border": border,
                "fill_header": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        }


def _write_header(sheet, period, user, styles):
        _write_centered(sheet, "A1", "H", "RELATÓRIO MENSAL DE FREQUÊNCIA", styles["title"])
        _write_centered(sheet, "A2", "H", f"Período: {period}", styles["value"])
        for info in _employee_info(user):
                _write_info_row(sheet, info, styles)


def _write_centered(sheet, cell, end_column, text, font):
        sheet.merge_cells(f"{cell}:{end_column}{cell[1]}")
        sheet[cell] = text
        sheet[cell].font = font
        sheet[cell].alignment = Alignment(horizontal="center")


def _employee_info(user):
        profile = user.profile
        return [
                ("A3", "Funcionário:", "B3", profile.full_name or "Não informado", "E3", "Matrícula:", "F3", profile.registration_number or "-"),
                ("A4", "Cargo:", "B4", profile.position or "-", "E4", "Departamento:", "F4", profile.department or "-"),
                ("A5", "CPF:", "B5", profile.cpf or "-", "E5", "Emitido em:", "F5", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]


def _write_info_row(sheet, info, styles):
        label_cell, label, value_cell, value, label2_cell, label2, value2_cell, value2 = info
        _write_cell(sheet, label_cell, label, styles["label"])
        _write_merged_value(sheet, value_cell, "D", value, styles["value"])
        _write_cell(sheet, label2_cell, label2, styles["label"])
        _write_merged_value(sheet, value2_cell, "H", value2, styles["value"])


def _write_merged_value(sheet, cell, end_column, text, font):
        sheet.merge_cells(f"{cell}:{end_column}{cell[1]}")
        sheet[cell] = text
        sheet[cell].font = font


def _write_cell(sheet, cell, text, font):
        sheet[cell] = text
        sheet[cell].font = font


def _style_table(sheet, row_count, styles):
        for cell in sheet[8]:
                cell.font = styles["header"]
                cell.fill = styles["fill_header"]
                cell.border = styles["border"]
                cell.alignment = Alignment(horizontal="center")
        _style_data_rows(sheet, row_count, styles)
        _set_column_widths(sheet)


def _style_data_rows(sheet, row_count, styles):
        for row in sheet.iter_rows(min_row=9, max_row=8 + row_count, min_col=1, max_col=8):
                for cell in row:
                        cell.font = styles["value"]
                        cell.border = styles["border"]
                        cell.alignment = Alignment(horizontal="center")


def _set_column_widths(sheet):
        for index, width in enumerate(_COLUMN_WIDTHS):
                sheet.column_dimensions[get_column_letter(index + 1)].width = width


def _write_signatures(sheet, row_count):
        signature_row = 8 + row_count + 3
        _write_signature_block(sheet, "A", "C", signature_row, "Assinatura do Funcionário")
        _write_signature_block(sheet, "F", "H", signature_row, "Assinatura do Gestor / RH")


def _write_signature_block(sheet, start_column, end_column, row, label):
        line_cell = f"{start_column}{row}"
        sheet.merge_cells(f"{line_cell}:{end_column}{row}")
        sheet[line_cell].border = Border(top=Side(style="thin"))
        label_cell = f"{start_column}{row + 1}"
        sheet.merge_cells(f"{label_cell}:{end_column}{row + 1}")
        sheet[label_cell] = label
        sheet[label_cell].font = Font(name="Arial", size=9)
        sheet[label_cell].alignment = Alignment(horizontal="center")
