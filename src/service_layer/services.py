from datetime import datetime, date, time, timedelta
from typing import Optional, List, Dict
from src.domain.model import User, UserRole, UserProfile, DailyPonto, Vacation, Holiday, WorkSchedule, PontoStatus, JourneyType, Notification, AuditLog, CorrectionRequest
from src.service_layer.unit_of_work import AbstractUnitOfWork
from werkzeug.security import generate_password_hash
import pandas as pd
import io

def ensure_manager(uow: AbstractUnitOfWork, manager_id: int):
    user = uow.users.get_user_by_id(manager_id)
    if not user or user.role not in [UserRole.MANAGER, UserRole.ADMIN]:
        raise PermissionError("Action restricted to managers or admins.")
    return user

def ensure_not_self(uow: AbstractUnitOfWork, manager_id: int, employee_id: int):
    if manager_id == employee_id:
        user = uow.users.get_user_by_id(manager_id)
        if not user or user.role != UserRole.ADMIN:
            raise PermissionError("Reviewers cannot review or correct their own time logs. This must be done by an Admin.")

def add_notification(uow: AbstractUnitOfWork, user_id: int, message: str, email_sender=None):
    notification = Notification(user_id=user_id, message=message, created_at=datetime.now())
    uow.session.add(notification)
    
    # Send email if enabled
    user = uow.users.get_user_by_id(user_id)
    if user and user.email_notifications_enabled and email_sender:
        email_sender(user.email, "Nova Notificação - Banco de Horas", f"<p>{message}</p>")

def mark_notifications_as_read(uow: AbstractUnitOfWork, user_id: int):
    with uow:
        user = uow.users.get_user_by_id(user_id)
        if user:
            # Deleta as notificações lidas permanentemente
            for n in user.notifications:
                uow.session.delete(n)
            uow.commit()

def register_user(
    uow: AbstractUnitOfWork, 
    email: str, 
    password: Optional[str] = None, 
    role: str = "employee",
    registered_by_id: Optional[int] = None
) -> bool:
    with uow:
        existing_user = uow.users.get_user_by_email(email)
        if existing_user:
            raise ValueError("Email already exists.")
        
        # If no password provided, use a placeholder
        pw_hash = generate_password_hash(password) if password else "!"
        
        user = User(
            email=email,
            password_hash=pw_hash,
            role=UserRole(role)
        )
        uow.users.add_user(user)
        uow.commit()
        
        actor_id = int(registered_by_id) if registered_by_id else user.user_id
        uow.record_action(actor_id, "USER_REGISTERED", target_id=user.user_id, details=f"Nível: {role}")
        uow.commit()
        return True

def update_user_profile(
    uow: AbstractUnitOfWork,
    user_id: int,
    registration_number: str,
    cpf: str,
    department: str,
    position: str,
    secretariat: str,
    full_name: str,
    start_analysis_date: Optional[date] = None
) -> None:
    with uow:
        user = uow.users.get_user_by_id(user_id)
        if not user:
            raise ValueError("User not found.")
        
        analysis_date = start_analysis_date or (user.profile.start_analysis_date if user.profile else None) or date(2026, 1, 1)
        
        new_profile = UserProfile(
            registration_number=registration_number,
            cpf=cpf,
            department=department,
            position=position,
            secretariat=secretariat,
            full_name=full_name,
            start_analysis_date=analysis_date
        )
        user.profile = new_profile
        uow.commit()
        uow.record_action(user_id, "UPDATE_PROFILE", target_id=user_id, details=f"Matrícula: {registration_number}, Depto: {department}")
        uow.commit()

def update_credentials(uow: AbstractUnitOfWork, user_id: int, email: str, password: Optional[str] = None, email_notifications_enabled: bool = False):
    with uow:
        user = uow.users.get_user_by_id(user_id)
        if not user:
            raise ValueError("User not found.")
        
        existing = uow.users.get_user_by_email(email)
        if existing and existing.user_id != user_id:
            raise ValueError("Email already in use.")
            
        user.email = email
        user.email_notifications_enabled = email_notifications_enabled
        if password:
            user.password_hash = generate_password_hash(password)
        uow.commit()
        uow.record_action(user_id, "UPDATE_CREDENTIALS", target_id=user_id)
        uow.commit()

def promote_to_manager(uow: AbstractUnitOfWork, manager_id: int, employee_id: int):
    with uow:
        ensure_manager(uow, manager_id)
        employee = uow.users.get_user_by_id(employee_id)
        if employee:
            employee.role = UserRole.MANAGER
            uow.commit()
            uow.record_action(manager_id, "PROMOTE_USER", target_id=employee_id, details="Promovido a Diretor")
            uow.commit()

def demote_to_employee(uow: AbstractUnitOfWork, manager_id: int, employee_id: int):
    with uow:
        ensure_manager(uow, manager_id)
        employee = uow.users.get_user_by_id(employee_id)
        if employee:
            employee.role = UserRole.EMPLOYEE
            uow.commit()
            uow.record_action(manager_id, "DEMOTE_USER", target_id=employee_id, details="Rebaixado a Funcionário")
            uow.commit()

import math

def calculate_distance(lat1, lon1, lat2, lon2):
    # Haversine formula
    R = 6371  # Earth radius in km
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat / 2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c

from src.domain.model import CompanySettings

def get_company_settings(uow: AbstractUnitOfWork) -> Optional[CompanySettings]:
    with uow:
        setting = uow.session.query(CompanySettings).first()
        return setting

def clock_in_out(uow: AbstractUnitOfWork, user_id: int, location: Optional[str] = None, stage: Optional[str] = None, notes: Optional[str] = None) -> str:
    loc = location or "Não obtida"

    with uow:
        user = uow.users.get_user_by_id(user_id)
        if not user:
            raise ValueError("User not found.")
        
        # Calculate Brasília Time (UTC-3)
        brazil_time = datetime.utcnow() - timedelta(hours=3)
        today = brazil_time.date()
        now_time = brazil_time.time()

        ponto = next((p for p in user.time_entries if p.entry_date == today), None)
        if ponto and ponto.location_data is None:
            ponto.location_data = ""

        if not stage:
            if not ponto: stage = "arrival"
            elif ponto.has_lunch_break and not ponto.lunch_start: stage = "lunch_start"
            elif ponto.has_lunch_break and not ponto.lunch_end: stage = "lunch_end"
            elif not ponto.departure: stage = "departure"
            else: raise ValueError("Jornada de hoje já está completa.")

        if stage == "arrival":
            if not ponto:
                has_lunch = user.work_schedule.has_lunch_break if user.work_schedule else True
                ponto = DailyPonto(user_id=user_id, entry_date=today, arrival=now_time, has_lunch_break=has_lunch)
                user.time_entries.append(ponto)
            else:
                ponto.arrival = now_time
            ponto.location_data += f" | Chegada: {loc}"
            msg = "Chegada registrada"
            if user.work_schedule:
                limit = (datetime.combine(today, user.work_schedule.expected_arrival) + 
                         timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                if now_time > limit:
                    ponto.status = PontoStatus.LATE
                    ponto.arrival_late = True
        elif stage == "lunch_start":
            if not ponto: raise ValueError("Registro de chegada não encontrado.")
            ponto.lunch_start = now_time
            ponto.location_data += f" | Almoço (Sai): {loc}"
            msg = "Saída para almoço registrada"
            if user.work_schedule and user.work_schedule.expected_lunch_start:
                limit = (datetime.combine(today, user.work_schedule.expected_lunch_start) - 
                         timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                if now_time < limit:
                    ponto.status = PontoStatus.LATE
                    ponto.lunch_start_late = True
        elif stage == "lunch_end":
            if not ponto: raise ValueError("Registro de chegada não encontrado.")
            ponto.lunch_end = now_time
            ponto.location_data += f" | Almoço (Vol): {loc}"
            msg = "Retorno do almoço registrado"
            if user.work_schedule and user.work_schedule.expected_lunch_end:
                limit = (datetime.combine(today, user.work_schedule.expected_lunch_end) + 
                         timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                if now_time > limit:
                    ponto.status = PontoStatus.LATE
                    ponto.lunch_end_late = True
        elif stage == "departure":
            if not ponto: raise ValueError("Registro de chegada não encontrado.")
            ponto.departure = now_time
            ponto.location_data += f" | Fim: {loc}"
            msg = "Fim de jornada registrado"
            if user.work_schedule:
                limit = (datetime.combine(today, user.work_schedule.expected_departure) - 
                         timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                if now_time < limit:
                    ponto.status = PontoStatus.LATE
                    ponto.departure_early = True
            if notes:
                ponto.notes = notes
        else:
            raise ValueError("Estágio inválido.")
        
        uow.commit()
        uow.record_action(user_id, "CLOCK_EVENT", target_id=user_id, details=f"{msg} ({stage})")
        uow.commit()
        return msg

def submit_correction_request(uow: AbstractUnitOfWork, user_id: int, ponto_date: date, stage: str, proposed_time: time):
    with uow:
        req = CorrectionRequest(
            user_id=user_id,
            ponto_date=ponto_date,
            stage=stage,
            proposed_time=proposed_time
        )
        uow.session.add(req)
        uow.commit()
        uow.record_action(user_id, "SUBMIT_CORRECTION_REQUEST", target_id=None, details=f"Data: {ponto_date}, Estágio: {stage}")
        uow.commit()

def list_pending_corrections(uow: AbstractUnitOfWork, manager_id: int) -> List[CorrectionRequest]:
    with uow:
        ensure_manager(uow, manager_id)
        return uow.session.query(CorrectionRequest).filter_by(status="pending").all()

def review_correction_request(uow: AbstractUnitOfWork, manager_id: int, request_id: int, approved: bool):
    with uow:
        ensure_manager(uow, manager_id)
        req = uow.session.query(CorrectionRequest).filter_by(id=request_id).first()
        if not req:
            raise ValueError("Solicitação não encontrada.")
        
        if approved:
            req.status = "approved"
            user = uow.users.get_user_by_id(req.user_id)
            ponto = next((p for p in user.time_entries if p.entry_date == req.ponto_date), None)
            if not ponto:
                ponto = DailyPonto(user_id=req.user_id, entry_date=req.ponto_date)
                user.time_entries.append(ponto)
            
            if ponto.location_data is None:
                ponto.location_data = ""
            
            if req.stage == "arrival": ponto.arrival = req.proposed_time
            elif req.stage == "lunch_start": ponto.lunch_start = req.proposed_time
            elif req.stage == "lunch_end": ponto.lunch_end = req.proposed_time
            elif req.stage == "departure": ponto.departure = req.proposed_time
            
            # Correction by manager implies review
            ponto.arrival_late_reviewed = True
            ponto.lunch_start_late_reviewed = True
            ponto.lunch_end_late_reviewed = True
            ponto.departure_early_reviewed = True
            
            # Re-evaluate anomalies after correction
            if user.work_schedule:
                # Arrival
                if ponto.arrival:
                    limit = (datetime.combine(req.ponto_date, user.work_schedule.expected_arrival) +
                             timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                    ponto.arrival_late = ponto.arrival > limit
                # Lunch Start
                if ponto.lunch_start and user.work_schedule.expected_lunch_start:
                    limit = (datetime.combine(req.ponto_date, user.work_schedule.expected_lunch_start) -
                             timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                    ponto.lunch_start_late = ponto.lunch_start < limit
                # Lunch End
                if ponto.lunch_end and user.work_schedule.expected_lunch_end:
                    limit = (datetime.combine(req.ponto_date, user.work_schedule.expected_lunch_end) +
                             timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                    ponto.lunch_end_late = ponto.lunch_end > limit
                # Departure
                if ponto.departure and user.work_schedule.expected_departure:
                    limit = (datetime.combine(req.ponto_date, user.work_schedule.expected_departure) -
                             timedelta(minutes=user.work_schedule.tolerance_minutes)).time()
                    ponto.departure_early = ponto.departure < limit

            ponto.status = PontoStatus.CORRECTED
            ponto.location_data += f" | Corrigido via solicitação aprovada por gestor {manager_id}"
            
            add_notification(uow, req.user_id, f"Sua solicitação de correção para {req.ponto_date} foi APROVADA.")
        else:
            req.status = "rejected"
            add_notification(uow, req.user_id, f"Sua solicitação de correção para {req.ponto_date} foi REJEITADA.")
        
        uow.commit()
        uow.record_action(manager_id, "REVIEW_CORRECTION_REQUEST", target_id=req.user_id, details=f"ID Pedido: {request_id}, Aprovado: {approved}")
        uow.commit()

def set_work_schedule(
    uow: AbstractUnitOfWork,
    manager_id: int,
    employee_id: int,
    arrival: time,
    lunch_start: Optional[time],
    lunch_end: Optional[time],
    departure: time,
    tolerance: int = 15,
    has_lunch_break: bool = True,
    schedule_type: str = "standard",
    rotation_start_date: Optional[date] = None
):
    with uow:
        user = uow.users.get_user_by_id(employee_id)
        if not user:
            raise ValueError("Employee not found.")

        # Allow self-assignment ONLY if no schedule exists
        if manager_id == employee_id:
            if user.work_schedule:
                raise PermissionError("Self-reassignment not allowed. Contact a manager.")
        else:
            ensure_manager(uow, manager_id)
        
        from src.domain.model import ScheduleType
        s_type = ScheduleType(schedule_type.lower())

        if user.work_schedule:
            user.work_schedule.expected_arrival = arrival
            user.work_schedule.expected_lunch_start = lunch_start
            user.work_schedule.expected_lunch_end = lunch_end
            user.work_schedule.expected_departure = departure
            user.work_schedule.tolerance_minutes = tolerance
            user.work_schedule.has_lunch_break = has_lunch_break
            user.work_schedule.schedule_type = s_type
            user.work_schedule.rotation_start_date = rotation_start_date
            uow.session.add(user.work_schedule)
        else:
            schedule = WorkSchedule(
                user_id=employee_id,
                expected_arrival=arrival,
                expected_lunch_start=lunch_start,
                expected_lunch_end=lunch_end,
                expected_departure=departure,
                tolerance_minutes=tolerance,
                has_lunch_break=has_lunch_break,
                schedule_type=s_type,
                rotation_start_date=rotation_start_date
            )
            user.work_schedule = schedule
            uow.session.add(schedule)
        
        uow.commit()

def manual_ponto_correction(
    uow: AbstractUnitOfWork, 
    manager_id: int, 
    employee_id: int, 
    entry_date: date,
    arrival: Optional[time],
    lunch_start: Optional[time],
    lunch_end: Optional[time],
    departure: Optional[time],
    manager_notes: Optional[str] = None,
    email_sender=None
) -> bool:
    with uow:
        manager = ensure_manager(uow, manager_id)
        ensure_not_self(uow, manager_id, employee_id)
        user = uow.users.get_user_by_id(employee_id)
        if not user:
            raise ValueError("Employee not found.")
        ponto = next((p for p in user.time_entries if p.entry_date == entry_date), None)
        
        if not ponto:
            if not any([arrival, lunch_start, lunch_end, departure]): return False
            if not ponto:
                has_lunch = user.work_schedule.has_lunch_break if user.work_schedule else True
                ponto = DailyPonto(user_id=employee_id, entry_date=entry_date, has_lunch_break=has_lunch)
                user.time_entries.append(ponto)

        changed = False
        if arrival is not None and ponto.arrival != arrival:
            ponto.arrival = arrival; changed = True
        if lunch_start is not None and ponto.lunch_start != lunch_start:
            ponto.lunch_start = lunch_start; changed = True
        if lunch_end is not None and ponto.lunch_end != lunch_end:
            ponto.lunch_end = lunch_end; changed = True
        if departure is not None and ponto.departure != departure:
            ponto.departure = departure; changed = True
        if manager_notes is not None and ponto.manager_notes != manager_notes:
            ponto.manager_notes = manager_notes
            changed = True
        
        if not changed: 
            return False

        ponto.status = PontoStatus.CORRECTED
        
        # Manual correction by manager implies review
        ponto.arrival_late_reviewed = True
        ponto.lunch_start_late_reviewed = True
        ponto.lunch_end_late_reviewed = True
        ponto.departure_early_reviewed = True

        if ponto.location_data is None:
            ponto.location_data = ""
        manager_name = manager.profile.full_name or manager.email
        ponto.location_data += f" | Corrigido manualmente por Gestor: {manager_name}"
        
        add_notification(uow, employee_id, f"Seu ponto de {entry_date} foi corrigido manualmente pelo gestor {manager_name}.", email_sender=email_sender)
        uow.session.flush()
        uow.commit()
        uow.record_action(manager_id, "MANUAL_CORRECTION", target_id=employee_id, details=f"Data: {entry_date}")
        uow.commit()
        return True

def generate_excel_report(uow: AbstractUnitOfWork, user_id: int, start_date: Optional[date] = None, end_date: Optional[date] = None) -> io.BytesIO:
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    with uow:
        user = uow.users.get_user_by_id(user_id)
        if not user:
            raise ValueError("Usuário não encontrado.")

        # Filter and Sort entries
        entries = user.time_entries
        if start_date:
            entries = [p for p in entries if p.entry_date >= start_date]
        if end_date:
            entries = [p for p in entries if p.entry_date <= end_date]
        
        entries = sorted(entries, key=lambda x: x.entry_date, reverse=False)

        # Header Info
        full_name = user.profile.full_name or "Não informado"
        registration = user.profile.registration_number or "-"
        cpf = user.profile.cpf or "-"
        department = user.profile.department or "-"
        position = user.profile.position or "-"

        # Period String
        period_str = "Todo o período"
        if start_date and end_date:
            period_str = f"{start_date.strftime('%d/%m/%Y')} até {end_date.strftime('%d/%m/%Y')}"
        elif start_date:
            period_str = f"A partir de {start_date.strftime('%d/%m/%Y')}"
        elif end_date:
            period_str = f"Até {end_date.strftime('%d/%m/%Y')}"

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Create a simple DataFrame for the core data
            data = []
            for p in entries:
                data.append({
                    "Data": p.entry_date.strftime("%d/%m/%Y"),
                    "Chegada": p.arrival.strftime("%H:%M:%S") if p.arrival else "-",
                    "Saída Almoço": p.lunch_start.strftime("%H:%M:%S") if p.lunch_start else "-",
                    "Volta Almoço": p.lunch_end.strftime("%H:%M:%S") if p.lunch_end else "-",
                    "Fim Jornada": p.departure.strftime("%H:%M:%S") if p.departure else "-",
                    "Horas Trab.": f"{p.worked_minutes // 60:02d}:{p.worked_minutes % 60:02d}",
                    "Status": p.status.value,
                    "Observações": p.notes or ""
                })

            df = pd.DataFrame(data)
            df.to_excel(writer, index=False, sheet_name="Relatório de Ponto", startrow=7)

            workbook = writer.book
            sheet = workbook["Relatório de Ponto"]

            # Styles
            title_font = Font(name='Arial', size=14, bold=True)
            header_font = Font(name='Arial', size=10, bold=True)
            label_font = Font(name='Arial', size=9, bold=True)
            value_font = Font(name='Arial', size=9)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            fill_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            # Report Title
            sheet.merge_cells('A1:H1')
            sheet['A1'] = "RELATÓRIO MENSAL DE FREQUÊNCIA"
            sheet['A1'].font = title_font
            sheet['A1'].alignment = Alignment(horizontal='center')

            sheet.merge_cells('A2:H2')
            sheet['A2'] = f"Período: {period_str}"
            sheet['A2'].font = value_font
            sheet['A2'].alignment = Alignment(horizontal='center')

            # Employee Info Section
            info_rows = [
                ('A3', 'Funcionário:', 'B3', full_name, 'E3', 'Matrícula:', 'F3', registration),
                ('A4', 'Cargo:', 'B4', position, 'E4', 'Departamento:', 'F4', department),
                ('A5', 'CPF:', 'B5', cpf, 'E5', 'Emitido em:', 'F5', datetime.now().strftime("%d/%m/%Y %H:%M"))
            ]

            for r in info_rows:
                sheet[r[0]] = r[1]; sheet[r[0]].font = label_font
                sheet.merge_cells(f"{r[2]}:D{r[0][1]}")
                sheet[r[2]] = r[3]; sheet[r[2]].font = value_font
                sheet[r[4]] = r[5]; sheet[r[4]].font = label_font
                sheet.merge_cells(f"{r[6]}:H{r[0][1]}")
                sheet[r[6]] = r[7]; sheet[r[6]].font = value_font

            # Style Table Headers
            for cell in sheet[8]:
                cell.font = header_font
                cell.fill = fill_header
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

            # Style Table Data
            last_row = 8 + len(data)
            for row in sheet.iter_rows(min_row=9, max_row=last_row, min_col=1, max_col=8):
                for cell in row:
                    cell.font = value_font
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

            # Adjust Column Widths
            widths = [12, 12, 12, 12, 12, 12, 15, 30]
            for i, width in enumerate(widths):
                sheet.column_dimensions[get_column_letter(i+1)].width = width

            # Signature Lines
            sig_row = last_row + 3
            sheet.merge_cells(f'A{sig_row}:C{sig_row}')
            sheet[f'A{sig_row}'].border = Border(top=Side(style='thin'))
            sheet[f'A{sig_row+1}'] = "Assinatura do Funcionário"
            sheet[f'A{sig_row+1}'].font = value_font
            sheet[f'A{sig_row+1}'].alignment = Alignment(horizontal='center')
            sheet.merge_cells(f'A{sig_row+1}:C{sig_row+1}')

            sheet.merge_cells(f'F{sig_row}:H{sig_row}')
            sheet[f'F{sig_row}'].border = Border(top=Side(style='thin'))
            sheet[f'F{sig_row+1}'] = "Assinatura do Gestor / RH"
            sheet[f'F{sig_row+1}'].font = value_font
            sheet[f'F{sig_row+1}'].alignment = Alignment(horizontal='center')
            sheet.merge_cells(f'F{sig_row+1}:H{sig_row+1}')

        output.seek(0)
        return output
def add_vacation(uow: AbstractUnitOfWork, manager_id: int, employee_id: int, start_date: date, end_date: date):
    with uow:
        ensure_manager(uow, manager_id)
        vacation = Vacation(user_id=employee_id, start_date=start_date, end_date=end_date)
        employee = uow.users.get_user_by_id(employee_id)
        if not employee:
            raise ValueError("Employee not found.")
        employee.vacations.append(vacation)
        uow.commit()
        uow.record_action(manager_id, "ADD_VACATION", target_id=employee_id, details=f"Início: {start_date}, Fim: {end_date}")
        uow.commit()

def delete_ponto_entry(uow: AbstractUnitOfWork, manager_id: int, employee_id: int, entry_date: date):
    with uow:
        manager = ensure_manager(uow, manager_id)
        ensure_not_self(uow, manager_id, employee_id)
        user = uow.users.get_user_by_id(employee_id)
        if not user:
            raise ValueError("Employee not found.")

        # Find all matching entries for this date
        matching_pontos = [p for p in user.time_entries if p.entry_date == entry_date]

        if matching_pontos:
            for ponto in matching_pontos:
                user.time_entries.remove(ponto)
                uow.session.delete(ponto)
            
            uow.commit()
            manager_name = manager.profile.full_name or manager.email
            uow.record_action(manager_id, "DELETE_PONTO", target_id=employee_id, details=f"Excluído(s) {len(matching_pontos)} registro(s) para {entry_date} por {manager_name}")
            uow.commit()
def add_holiday(uow: AbstractUnitOfWork, manager_id: int, holiday_date: date, description: str, is_mandatory: bool = True):
    with uow:
        ensure_manager(uow, manager_id)
        holiday = Holiday(holiday_date=holiday_date, description=description, is_mandatory=is_mandatory)
        uow.session.merge(holiday)
        uow.commit()
        uow.record_action(manager_id, "ADD_HOLIDAY", target_id=None, details=f"Data: {holiday_date}, Desc: {description}")
        uow.commit()

def get_start_analysis_date(uow: AbstractUnitOfWork) -> date:
    with uow:
        settings = uow.session.query(CompanySettings).first()
        return settings.start_analysis_date if settings else date(2026, 1, 1)

def get_all_employees(uow: AbstractUnitOfWork, requester_id: Optional[int] = None) -> List[User]:
    with uow:
        if requester_id:
            user = uow.users.get_user_by_id(requester_id)
            if user and user.role == UserRole.ADMIN:
                return uow.users.list_all()
        return uow.users.list_employees()

def delete_user(uow: AbstractUnitOfWork, manager_id: int, user_id: int):
    with uow:
        ensure_manager(uow, manager_id)
        user = uow.users.get_user_by_id(user_id)
        if user:
            # Preserve employee name in audit logs before nullifying user_id
            employee_name = user.profile.full_name if user.profile and user.profile.full_name else user.email
            # Update existing audit log entries to include employee name in details
            audit_entries = uow.session.query(AuditLog).filter_by(user_id=user_id).all()
            for audit in audit_entries:
                existing = audit.details or ""
                audit.details = f"{existing} (User: {employee_name})" if existing else f"User: {employee_name}"
            uow.session.flush()
            # Nullify user_id in audit logs for this user (to avoid FK constraints)
            uow.session.query(AuditLog).filter_by(user_id=user_id).update({AuditLog.user_id: None})
            
            email = user.email
            uow.session.delete(user)
            uow.commit()
            uow.record_action(manager_id, "DELETE_USER", target_id=user_id, details=f"Usuário deletado: {email}")
            uow.commit()

def create_journey_type(
    uow: AbstractUnitOfWork,
    manager_id: int,
    name: str,
    arrival: time,
    lunch_start: Optional[time],
    lunch_end: Optional[time],
    departure: time,
    tolerance: int = 15,
    has_lunch_break: bool = True,
    schedule_type: str = "standard"
):
    with uow:
        ensure_manager(uow, manager_id)
        from src.domain.model import ScheduleType
        jt = JourneyType(
            name=name,
            expected_arrival=arrival,
            expected_lunch_start=lunch_start,
            expected_lunch_end=lunch_end,
            expected_departure=departure,
            tolerance_minutes=tolerance,
            has_lunch_break=has_lunch_break,
            schedule_type=ScheduleType(schedule_type)
        )
        uow.session.add(jt)
        uow.commit()
        uow.record_action(manager_id, "CREATE_JOURNEY_TYPE", target_id=None, details=f"Nome: {name}, Almoço: {has_lunch_break}, Tipo: {schedule_type}")
        uow.commit()

def list_journey_types(uow: AbstractUnitOfWork) -> List[JourneyType]:
    with uow:
        return uow.session.query(JourneyType).all()

def get_journey_type(uow: AbstractUnitOfWork, journey_id: int) -> Optional[JourneyType]:
    with uow:
        return uow.session.query(JourneyType).filter_by(journey_id=journey_id).first()

def update_journey_type(
    uow: AbstractUnitOfWork,
    manager_id: int,
    journey_id: int,
    name: str,
    arrival: time,
    lunch_start: Optional[time],
    lunch_end: Optional[time],
    departure: time,
    tolerance: int = 15,
    has_lunch_break: bool = True,
    schedule_type: str = "standard"
):
    with uow:
        ensure_manager(uow, manager_id)
        jt = uow.session.query(JourneyType).filter_by(journey_id=journey_id).first()
        if not jt:
            raise ValueError("Journey Type not found.")
        
        from src.domain.model import ScheduleType
        jt.name = name
        jt.expected_arrival = arrival
        jt.expected_lunch_start = lunch_start
        jt.expected_lunch_end = lunch_end
        jt.expected_departure = departure
        jt.tolerance_minutes = tolerance
        jt.has_lunch_break = has_lunch_break
        jt.schedule_type = ScheduleType(schedule_type)
        uow.commit()
        uow.record_action(manager_id, "UPDATE_JOURNEY_TYPE", target_id=journey_id, details=f"Nome: {name}, Almoço: {has_lunch_break}, Tipo: {schedule_type}")
        uow.commit()

def delete_journey_type(uow: AbstractUnitOfWork, manager_id: int, journey_id: int):
    with uow:
        ensure_manager(uow, manager_id)
        jt = uow.session.query(JourneyType).filter_by(journey_id=journey_id).first()
        if jt:
            name = jt.name
            uow.session.delete(jt)
            uow.commit()
            uow.record_action(manager_id, "DELETE_JOURNEY_TYPE", target_id=journey_id, details=f"Deletado: {name}")
            uow.commit()

def review_anomaly_badge(uow: AbstractUnitOfWork, admin_id: int, employee_id: int, entry_date: date, stage: str, action: str):
    with uow:
        admin = uow.users.get_user_by_id(admin_id)
        if not admin or admin.role != UserRole.ADMIN:
            raise PermissionError("Apenas o Administrador pode aprovar anomalias individuais.")
        
        user = uow.users.get_user_by_id(employee_id)
        if not user: raise ValueError("Employee not found.")
        
        ponto = next((p for p in user.time_entries if p.entry_date == entry_date), None)
        if not ponto: raise ValueError("Registro não encontrado.")
        
        if stage == "arrival": 
            ponto.arrival_late_reviewed = True
            if action == "approve": ponto.arrival_late_approved = True
            elif action == "excuse": ponto.arrival_late_excused = True
        elif stage == "lunch_start": 
            ponto.lunch_start_late_reviewed = True
            if action == "approve": ponto.lunch_start_late_approved = True
            elif action == "excuse": ponto.lunch_start_late_excused = True
        elif stage == "lunch_end": 
            ponto.lunch_end_late_reviewed = True
            if action == "approve": ponto.lunch_end_late_approved = True
            elif action == "excuse": ponto.lunch_end_late_excused = True
        elif stage == "departure": 
            ponto.departure_early_reviewed = True
            if action == "approve": ponto.departure_early_approved = True
            elif action == "excuse": ponto.departure_early_excused = True
        
        uow.commit()
        uow.record_action(admin_id, f"{action.upper()}_ANOMALY", target_id=employee_id, details=f"Data: {entry_date}, Estágio: {stage}")
        uow.commit()
